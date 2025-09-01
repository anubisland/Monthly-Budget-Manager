#!/usr/bin/env python3
"""
Budget Manager GUI (Tkinter)
- Manage incomes and expenses with per-entry date (YYYY-MM or YYYY-MM-DD)
- CSV import/export compatible with CLI; includes 'date' column
- Live report totals, profit margin, and category percentages
Run: python budget_manager_gui.py
"""
from __future__ import annotations

import csv
import json
import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
import datetime as _dt
import calendar as _cal
import math as _math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, numbers
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side

from budget_manager import BudgetMonth, read_csv


def _valid_ym(s: str) -> bool:
    return len(s) == 7 and s[4] == '-' and s[:4].isdigit() and s[5:7].isdigit() and 1 <= int(s[5:7]) <= 12


def _valid_ymd(s: str) -> bool:
    if len(s) != 10 or s[4] != '-' or s[7] != '-':
        return False
    y, m, d = s.split('-')
    if not (y.isdigit() and m.isdigit() and d.isdigit()):
        return False
    try:
        _dt.date(int(y), int(m), int(d))
        return True
    except ValueError:
        return False


class BudgetApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Budget Manager")
        self.geometry("900x600")

        self.bm = BudgetMonth()

        # Internal state
        self._chart_redraw_after = None
        self._prefs_path = None

        # Tk variables
        self.var_month = tk.StringVar()

        self.var_income_name = tk.StringVar()
        self.var_income_amount = tk.StringVar()
        self.var_income_date = tk.StringVar()

        self.var_expense_name = tk.StringVar()
        self.var_expense_category = tk.StringVar()
        self.var_expense_amount = tk.StringVar()
        self.var_expense_date = tk.StringVar()

        # Preferences path (user home)
        try:
            self._prefs_path = Path.home() / ".budget_manager_prefs.json"
        except Exception:
            self._prefs_path = None

        self._build_menu()
        self._build_toolbar()
        self._build_body()
        self._build_statusbar()
        # Load preferences after widgets exist
        self._load_prefs()
        self.update_report()

    def _build_menu(self) -> None:
        m = tk.Menu(self)
        filem = tk.Menu(m, tearoff=0)
        filem.add_command(label="New Budget", command=self.new_budget)
        filem.add_separator()
        filem.add_command(label="Open CSV…", command=self.open_csv, accelerator="Ctrl+O")
        filem.add_command(label="Save CSV As…", command=self.save_csv, accelerator="Ctrl+S")
        filem.add_separator()
        filem.add_command(label="Export JSON Report…", command=self.export_json)
        filem.add_separator()
        filem.add_command(label="Exit", command=self.quit)
        m.add_cascade(label="File", menu=filem)
        self.config(menu=m)
        # Global shortcuts
        self.bind_all("<Control-o>", lambda e: self.open_csv())
        self.bind_all("<Control-s>", lambda e: self.save_csv())
        self.bind_all("<F5>", lambda e: self.update_report())

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self, padding=(8, 6))
        ttk.Label(bar, text="Month (YYYY-MM)").grid(row=0, column=0, sticky="w")
        e = ttk.Entry(bar, textvariable=self.var_month, width=12)
        e.grid(row=0, column=1, padx=(6, 16))
        e.bind("<FocusOut>", lambda _e: self._on_month_changed())
        ttk.Button(bar, text="New", command=self.new_budget).grid(row=0, column=2, padx=4)
        ttk.Button(bar, text="Open CSV…", command=self.open_csv).grid(row=0, column=3, padx=4)
        ttk.Button(bar, text="Save CSV…", command=self.save_csv).grid(row=0, column=4, padx=4)
        ttk.Button(bar, text="Export Excel…", command=self.export_excel).grid(row=0, column=5, padx=4)
        ttk.Button(bar, text="Refresh Report", command=self.update_report).grid(row=0, column=6, padx=16)
        bar.columnconfigure(7, weight=1)
        bar.grid(row=0, column=0, sticky="ew")

    def _build_body(self) -> None:
        nb = ttk.Notebook(self)
        tab_income = ttk.Frame(nb, padding=8)
        tab_expense = ttk.Frame(nb, padding=8)
        tab_report = ttk.Frame(nb, padding=8)
        nb.add(tab_income, text="Income")
        nb.add(tab_expense, text="Expenses")
        nb.add(tab_report, text="Report")
        nb.grid(row=1, column=0, sticky="nsew")
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        # Income tab
        form = ttk.Frame(tab_income)
        ttk.Label(form, text="Name").grid(row=0, column=0, sticky="w")
        self.ent_income_name = ttk.Entry(form, textvariable=self.var_income_name, width=24)
        self.ent_income_name.grid(row=1, column=0, padx=(0, 8))
        ttk.Label(form, text="Amount").grid(row=0, column=1, sticky="w")
        self.ent_income_amount = ttk.Entry(form, textvariable=self.var_income_amount, width=12)
        self.ent_income_amount.grid(row=1, column=1, padx=(0, 8))
        ttk.Label(form, text="Date (YYYY-MM, YYYY-MM-DD, or DD)").grid(row=0, column=2, sticky="w")
        self.ent_income_date = ttk.Entry(form, textvariable=self.var_income_date, width=12)
        self.ent_income_date.grid(row=1, column=2, padx=(0, 4))
        ttk.Button(form, text="Pick…", command=self.pick_income_date).grid(row=1, column=3, padx=(0, 4))
        ttk.Button(form, text="Today", command=self.use_today_income_date).grid(row=1, column=4, padx=(0, 8))
        ttk.Button(form, text="Add Income", command=self.add_income).grid(row=1, column=5)
        form.grid(row=0, column=0, sticky="w", pady=(0, 8))
        for w in (self.ent_income_name, self.ent_income_amount, self.ent_income_date):
            w.bind("<Return>", lambda e: self.add_income())
            w.bind("<Escape>", lambda e: self._clear_income_fields())

        self.tv_income = ttk.Treeview(
            tab_income,
            columns=("name", "amount", "date"),
            show="headings",
            selectmode="extended",
        )
        self.tv_income.heading("name", text="Name")
        self.tv_income.heading("amount", text="Amount ($)")
        self.tv_income.heading("date", text="Day (weekday)")
        self.tv_income.column("name", width=300, anchor="w")
        self.tv_income.column("amount", width=120, anchor="e")
        self.tv_income.column("date", width=110, anchor="center")
        vsb = ttk.Scrollbar(tab_income, orient="vertical", command=self.tv_income.yview)
        self.tv_income.configure(yscrollcommand=vsb.set)
        self.tv_income.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")
        btns = ttk.Frame(tab_income)
        ttk.Button(btns, text="Remove Selected", command=self.remove_income_selected).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns, text="Clear All", command=self.clear_incomes).grid(row=0, column=1)
        btns.grid(row=2, column=0, pady=8, sticky="w")
        tab_income.columnconfigure(0, weight=1)
        tab_income.rowconfigure(1, weight=1)

        # Expense tab
        form2 = ttk.Frame(tab_expense)
        ttk.Label(form2, text="Name").grid(row=0, column=0, sticky="w")
        ttk.Entry(form2, textvariable=self.var_expense_name, width=24).grid(row=1, column=0, padx=(0, 8))
        ttk.Label(form2, text="Category").grid(row=0, column=1, sticky="w")
        cat_field = ttk.Frame(form2)
        self.ent_expense_category = ttk.Entry(cat_field, textvariable=self.var_expense_category, width=18)
        self.ent_expense_category.grid(row=0, column=0)
        ttk.Button(cat_field, text="Pick…", width=6, command=self.pick_category).grid(row=0, column=1, padx=(4, 0))
        cat_field.grid(row=1, column=1, padx=(0, 8), sticky="w")
        ttk.Label(form2, text="Amount").grid(row=0, column=2, sticky="w")
        self.ent_expense_amount = ttk.Entry(form2, textvariable=self.var_expense_amount, width=12)
        self.ent_expense_amount.grid(row=1, column=2, padx=(0, 8))
        ttk.Label(form2, text="Date (YYYY-MM, YYYY-MM-DD, or DD)").grid(row=0, column=3, sticky="w")
        self.ent_expense_date = ttk.Entry(form2, textvariable=self.var_expense_date, width=12)
        self.ent_expense_date.grid(row=1, column=3, padx=(0, 4))
        ttk.Button(form2, text="Pick…", command=self.pick_expense_date).grid(row=1, column=4, padx=(0, 4))
        ttk.Button(form2, text="Today", command=self.use_today_expense_date).grid(row=1, column=5, padx=(0, 8))
        ttk.Button(form2, text="Add Expense", command=self.add_expense).grid(row=1, column=6)
        form2.grid(row=0, column=0, sticky="w", pady=(0, 8))
        for w in (self.ent_expense_amount, self.ent_expense_date, self.ent_expense_category):
            w.bind("<Return>", lambda e: self.add_expense())
            w.bind("<Escape>", lambda e: self._clear_expense_fields())

        self.tv_expense = ttk.Treeview(
            tab_expense,
            columns=("name", "category", "amount", "date"),
            show="headings",
            selectmode="extended",
        )
        self.tv_expense.heading("name", text="Name")
        self.tv_expense.heading("category", text="Category")
        self.tv_expense.heading("amount", text="Amount ($)")
        self.tv_expense.heading("date", text="Day (weekday)")
        self.tv_expense.column("name", width=260, anchor="w")
        self.tv_expense.column("category", width=180, anchor="w")
        self.tv_expense.column("amount", width=120, anchor="e")
        self.tv_expense.column("date", width=110, anchor="center")
        vsb2 = ttk.Scrollbar(tab_expense, orient="vertical", command=self.tv_expense.yview)
        self.tv_expense.configure(yscrollcommand=vsb2.set)
        self.tv_expense.grid(row=1, column=0, sticky="nsew")
        vsb2.grid(row=1, column=1, sticky="ns")
        btns2 = ttk.Frame(tab_expense)
        ttk.Button(btns2, text="Remove Selected", command=self.remove_expense_selected).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(btns2, text="Clear All", command=self.clear_expenses).grid(row=0, column=1)
        btns2.grid(row=2, column=0, pady=8, sticky="w")
        tab_expense.columnconfigure(0, weight=1)
        tab_expense.rowconfigure(1, weight=1)

        # Report tab
        summary = ttk.LabelFrame(tab_report, text="Summary", padding=8)
        self.lbl_income = ttk.Label(summary, text="Total Income: $0.00")
        self.lbl_expenses = ttk.Label(summary, text="Total Expenses: $0.00")
        self.lbl_net = ttk.Label(summary, text="Net (Profit): $0.00")
        self.lbl_margin = ttk.Label(summary, text="Profit Margin: 0.00%")
        self.lbl_income.grid(row=0, column=0, sticky="w", padx=(0, 16))
        self.lbl_expenses.grid(row=0, column=1, sticky="w", padx=(0, 16))
        self.lbl_net.grid(row=0, column=2, sticky="w", padx=(0, 16))
        self.lbl_margin.grid(row=0, column=3, sticky="w")
        summary.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        charts = ttk.LabelFrame(tab_report, text="Charts", padding=8)
        charts.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        charts.columnconfigure(0, weight=1)
        charts.columnconfigure(1, weight=1)
        left = ttk.Frame(charts)
        right = ttk.Frame(charts)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        ttk.Label(left, text="Income vs Expenses").grid(row=0, column=0, sticky="w")
        self.canvas_income_expense = tk.Canvas(left, width=380, height=230, bg="white", highlightthickness=1, highlightbackground="#ddd")
        self.canvas_income_expense.grid(row=1, column=0, sticky="nsew")
        ttk.Label(right, text="Expense Categories").grid(row=0, column=0, sticky="w")
        self.canvas_categories = tk.Canvas(right, width=380, height=230, bg="white", highlightthickness=1, highlightbackground="#ddd")
        self.canvas_categories.grid(row=1, column=0, sticky="nsew")
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        self.canvas_income_expense.bind("<Configure>", lambda e: self._schedule_redraw())
        self.canvas_categories.bind("<Configure>", lambda e: self._schedule_redraw())

        self.tv_breakdown = ttk.Treeview(
            tab_report,
            columns=("category", "amount", "p_income", "p_expenses"),
            show="headings",
        )
        self.tv_breakdown.heading("category", text="Category")
        self.tv_breakdown.heading("amount", text="Amount ($)")
        self.tv_breakdown.heading("p_income", text="% of Income")
        self.tv_breakdown.heading("p_expenses", text="% of Expenses")
        self.tv_breakdown.column("category", anchor="w", width=260)
        self.tv_breakdown.column("amount", anchor="e", width=120)
        self.tv_breakdown.column("p_income", anchor="e", width=120)
        self.tv_breakdown.column("p_expenses", anchor="e", width=120)
        vsb3 = ttk.Scrollbar(tab_report, orient="vertical", command=self.tv_breakdown.yview)
        self.tv_breakdown.configure(yscrollcommand=vsb3.set)
        self.tv_breakdown.grid(row=2, column=0, sticky="nsew")
        vsb3.grid(row=2, column=1, sticky="ns")
        tab_report.columnconfigure(0, weight=1)
        tab_report.rowconfigure(2, weight=1)
    # (Column sorting can be added later)

    def _build_statusbar(self) -> None:
        self.status = tk.StringVar(value="Ready")
        bar = ttk.Frame(self, padding=(8, 4))
        ttk.Label(bar, textvariable=self.status, anchor="w").grid(row=0, column=0, sticky="ew")
        bar.columnconfigure(0, weight=1)
        bar.grid(row=2, column=0, sticky="ew")

    # Actions
    def _on_month_changed(self) -> None:
        self.bm.month = self.var_month.get().strip() or None
        self.update_report()
        self._save_prefs()

    def add_income(self) -> None:
        name = (self.var_income_name.get() or "Income").strip()
        amt_str = (self.var_income_amount.get() or "0").replace(",", "").strip()
        raw_date = (self.var_income_date.get() or "").strip()
        try:
            amt = float(amt_str)
            if amt < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid amount", "Income amount must be non-negative.")
            return
        norm_date, err = self._normalize_date_input(raw_date)
        if err:
            messagebox.showerror("Invalid date", err)
            return
        self.bm.add_income(name, amt, norm_date)
        self._clear_income_fields()
        self.refresh_income_view()
        self.update_report()
        self.status.set("Income added.")

    def add_expense(self) -> None:
        name = (self.var_expense_name.get() or "Expense").strip()
        category = (self.var_expense_category.get() or "Uncategorized").strip()
        amt_str = (self.var_expense_amount.get() or "0").replace(",", "").strip()
        raw_date = (self.var_expense_date.get() or "").strip()
        try:
            amt = float(amt_str)
            if amt < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid amount", "Expense amount must be non-negative.")
            return
        norm_date, err = self._normalize_date_input(raw_date)
        if err:
            messagebox.showerror("Invalid date", err)
            return
        self.bm.add_expense(name, amt, category, norm_date)
        self._clear_expense_fields()
        self.refresh_expense_view()
        self.update_report()
        self.status.set("Expense added.")

    def _clear_income_fields(self) -> None:
        self.var_income_name.set("")
        self.var_income_amount.set("")
        self.var_income_date.set("")

    def _clear_expense_fields(self) -> None:
        self.var_expense_name.set("")
        self.var_expense_category.set("")
        self.var_expense_amount.set("")
        self.var_expense_date.set("")

    def remove_income_selected(self) -> None:
        sel = list(self.tv_income.selection())
        if not sel:
            return
        indices = sorted((self.tv_income.index(i) for i in sel), reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.bm.incomes):
                del self.bm.incomes[idx]
        self.refresh_income_view()
        self.update_report()
        self.status.set("Income removed.")

    def remove_expense_selected(self) -> None:
        sel = list(self.tv_expense.selection())
        if not sel:
            return
        indices = sorted((self.tv_expense.index(i) for i in sel), reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.bm.expenses):
                del self.bm.expenses[idx]
        self.refresh_expense_view()
        self.update_report()
        self.status.set("Expense removed.")

    # Date helpers
    def _open_calendar(self, target_var: tk.StringVar) -> None:
        base = (self.var_month.get() or "").strip()
        today = _dt.date.today()
        year = today.year
        month = today.month
        if _valid_ym(base):
            year = int(base[:4])
            month = int(base[5:7])
        sel = _CalendarPopup(self, year, month).show()
        if sel:
            target_var.set(sel)

    def pick_income_date(self) -> None:
        self._open_calendar(self.var_income_date)

    def pick_expense_date(self) -> None:
        self._open_calendar(self.var_expense_date)

    def use_today_income_date(self) -> None:
        self.var_income_date.set(_dt.date.today().strftime("%Y-%m-%d"))

    def use_today_expense_date(self) -> None:
        self.var_expense_date.set(_dt.date.today().strftime("%Y-%m-%d"))

    def pick_category(self) -> None:
        cats = [
            "Food", "Rent", "Fuel", "Electricity", "Internet", "Water", "Transport",
            "Healthcare", "Entertainment", "Education", "Clothing", "Savings",
            "Debt", "Subscriptions", "Gifts", "Misc"
        ]
        sel = _CategoryPicker(self, cats).show()
        if sel:
            self.var_expense_category.set(sel)

    def clear_incomes(self) -> None:
        if messagebox.askyesno("Clear incomes", "Remove all income entries?"):
            self.bm.incomes.clear()
            self.refresh_income_view()
            self.update_report()
            self.status.set("All incomes cleared.")

    def clear_expenses(self) -> None:
        if messagebox.askyesno("Clear expenses", "Remove all expense entries?"):
            self.bm.expenses.clear()
            self.refresh_expense_view()
            self.update_report()
            self.status.set("All expenses cleared.")

    def new_budget(self) -> None:
        if not self._confirm_discard_changes():
            return
        self.bm = BudgetMonth()
        self.var_month.set("")
        self.refresh_income_view()
        self.refresh_expense_view()
        self.update_report()
        self.status.set("New budget created.")

    # File ops
    def open_csv(self) -> None:
        path = filedialog.askopenfilename(title="Open CSV", filetypes=[["CSV files", "*.csv"], ["All files", "*.*"]])
        if not path:
            return
        try:
            incomes, expenses = read_csv(Path(path))
        except Exception as exc:
            messagebox.showerror("Failed to open CSV", str(exc))
            return
        self.bm.incomes = list(incomes)
        self.bm.expenses = list(expenses)
        inferred = self._infer_month_from_entries()
        if inferred:
            self.var_month.set(inferred)
            self.bm.month = inferred
            self._save_prefs()
        self.refresh_income_view()
        self.refresh_expense_view()
        self.update_report()
        self.status.set(f"Loaded {os.path.basename(path)}")

    def save_csv(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save CSV",
            defaultextension=".csv",
            filetypes=[["CSV files", "*.csv"], ["All files", "*.*"]],
            initialfile="budget.csv",
        )
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["type", "name", "category", "amount", "date"])  # date = YYYY-MM or YYYY-MM-DD
                for inc in self.bm.incomes:
                    w.writerow(["income", inc.name, "", f"{inc.amount:.2f}", (inc.date or self.bm.month or "")])
                for exp in self.bm.expenses:
                    w.writerow(["expense", exp.name, exp.category, f"{exp.amount:.2f}", (exp.date or self.bm.month or "")])
        except Exception as exc:
            messagebox.showerror("Failed to save CSV", str(exc))
            return
        self.status.set(f"Saved to {os.path.basename(path)}")

    def export_json(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Export JSON Report",
            defaultextension=".json",
            filetypes=[["JSON files", "*.json"], ["All files", "*.*"]],
            initialfile="budget_report.json",
        )
        if not path:
            return
        try:
            data = self.bm.to_dict()
            data["month"] = self.var_month.get().strip() or self.bm.month
            Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")
        except Exception as exc:
            messagebox.showerror("Failed to export JSON", str(exc))
            return
        self.status.set(f"JSON report saved: {os.path.basename(path)}")

    def export_excel(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Export Excel Workbook",
            defaultextension=".xlsx",
            filetypes=[["Excel Workbook", "*.xlsx"], ["All files", "*.*"]],
            initialfile="budget_report.xlsx",
        )
        if not path:
            return
        try:
            wb = Workbook()
            # Income sheet
            ws_income = wb.active
            ws_income.title = "Income"
            ws_income.append(["Name", "Amount", "Date", "Day (weekday)"])
            for cell in ws_income[1]:
                cell.font = Font(bold=True)
            for inc in self.bm.incomes:
                ds = inc.date or (self.var_month.get().strip() or self.bm.month or "")
                ws_income.append([
                    inc.name,
                    float(inc.amount),
                    ds,
                    self._format_day_display(ds),
                ])
            data_end_income = ws_income.max_row
            # Totals row
            ws_income.append(["Total", f"=SUM(B2:B{data_end_income})", "", ""])
            tot_row_income = ws_income.max_row
            ws_income[f"A{tot_row_income}"].font = Font(bold=True)
            ws_income[f"B{tot_row_income}"].font = Font(bold=True)
            # Table style (without totals row)
            income_table = Table(displayName="IncomeTable", ref=f"A1:D{data_end_income}")
            income_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            ws_income.add_table(income_table)
            ws_income.column_dimensions['A'].width = 32
            ws_income.column_dimensions['B'].width = 14
            ws_income.column_dimensions['C'].width = 14
            ws_income.column_dimensions['D'].width = 16
            for row in ws_income.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    cell.alignment = Alignment(horizontal="right")
            # Emphasize totals row
            thin = Side(style="thin", color="999999")
            top = Side(style="medium", color="666666")
            fill_total = PatternFill("solid", fgColor="FFFBEA")
            for col in (1, 2, 3, 4):
                c = ws_income.cell(row=tot_row_income, column=col)
                c.fill = fill_total
                c.border = Border(top=top, left=thin, right=thin, bottom=thin)
            # Freeze header
            ws_income.freeze_panes = "A2"

            # Expenses sheet
            ws_exp = wb.create_sheet("Expenses")
            ws_exp.append(["Name", "Category", "Amount", "Date", "Day (weekday)"])
            for cell in ws_exp[1]:
                cell.font = Font(bold=True)
            for exp in self.bm.expenses:
                ds = exp.date or (self.var_month.get().strip() or self.bm.month or "")
                ws_exp.append([
                    exp.name,
                    exp.category,
                    float(exp.amount),
                    ds,
                    self._format_day_display(ds),
                ])
            data_end_exp = ws_exp.max_row
            ws_exp.append(["Total", "", f"=SUM(C2:C{data_end_exp})", "", ""]) 
            tot_row_exp = ws_exp.max_row
            ws_exp[f"A{tot_row_exp}"].font = Font(bold=True)
            ws_exp[f"C{tot_row_exp}"].font = Font(bold=True)
            # Table style (without totals row)
            exp_table = Table(displayName="ExpensesTable", ref=f"A1:E{data_end_exp}")
            exp_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
            ws_exp.add_table(exp_table)
            ws_exp.column_dimensions['A'].width = 30
            ws_exp.column_dimensions['B'].width = 20
            ws_exp.column_dimensions['C'].width = 14
            ws_exp.column_dimensions['D'].width = 14
            ws_exp.column_dimensions['E'].width = 16
            for row in ws_exp.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    cell.alignment = Alignment(horizontal="right")
            # Emphasize totals row
            for col in (1, 2, 3, 4, 5):
                c = ws_exp.cell(row=tot_row_exp, column=col)
                c.fill = fill_total
                c.border = Border(top=top, left=thin, right=thin, bottom=thin)
            # Freeze header
            ws_exp.freeze_panes = "A2"

            # Report sheet
            ws_rep = wb.create_sheet("Report")
            month = self.var_month.get().strip() or (self.bm.month or "")
            ws_rep.append(["Month", month])
            ws_rep.append(["Total Income", float(self.bm.total_income())])
            ws_rep.append(["Total Expenses", float(self.bm.total_expenses())])
            ws_rep.append(["Net (Profit)", float(self.bm.net())])
            ws_rep.append(["Profit Margin", f"{self.bm.profit_margin():.2f}%"]) 
            for r in range(1, 6):
                ws_rep[f"A{r}"].font = Font(bold=True)
            ws_rep.column_dimensions['A'].width = 20
            ws_rep.column_dimensions['B'].width = 20
            ws_rep["B2"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_rep["B3"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_rep["B4"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            ws_rep.append([])
            ws_rep.append(["Expense Breakdown by Category"]) 
            ws_rep.append(["Category", "Amount", "% of Income", "% of Expenses"]) 
            ws_rep[ws_rep.max_row - 0][0].font = Font(bold=True)
            for cell in ws_rep[ws_rep.max_row]:
                cell.font = Font(bold=True)
            by_cat = self.bm.expenses_by_category()
            p_inc = self.bm.expense_percentages_by_category("income")
            p_exp = self.bm.expense_percentages_by_category("expenses")
            start_row = ws_rep.max_row + 1
            for cat, amt in sorted(by_cat.items()):
                # Write numeric percentages (0-1) and apply % number format later
                ws_rep.append([cat, float(amt), (p_inc.get(cat, 0.0) / 100.0), (p_exp.get(cat, 0.0) / 100.0)]) 
            end_row = ws_rep.max_row
            # Amount formatting for the table
            for row in ws_rep.iter_rows(min_row=start_row, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    cell.alignment = Alignment(horizontal="right")
            # Percentage formatting
            for row in ws_rep.iter_rows(min_row=start_row, min_col=3, max_col=4):
                for cell in row:
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal="right")
            # Table style for breakdown (without totals)
            header_row = start_row - 1
            if end_row >= start_row:
                rep_table = Table(displayName="BreakdownTable", ref=f"A{header_row}:D{end_row}")
                rep_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
                ws_rep.add_table(rep_table)
            # Totals row for breakdown
            ws_rep.append(["Total", f"=SUM(B{start_row}:B{end_row})", "", ""]) 
            tot_row_rep = ws_rep.max_row
            ws_rep[f"A{tot_row_rep}"].font = Font(bold=True)
            ws_rep[f"B{tot_row_rep}"].font = Font(bold=True)
            for col in (1, 2, 3, 4):
                c = ws_rep.cell(row=tot_row_rep, column=col)
                c.fill = fill_total
                c.border = Border(top=top, left=thin, right=thin, bottom=thin)
            # Freeze header of breakdown section
            ws_rep.freeze_panes = f"A{start_row}"

            # Auto-fit columns for better readability
            self._excel_autofit_columns(ws_income)
            self._excel_autofit_columns(ws_exp)
            # Build charts on Report sheet
            # Bar: Income vs Expenses (B2, B3)
            bar = BarChart()
            bar.title = "Income vs Expenses"
            bar.style = 10
            data = Reference(ws_rep, min_col=2, min_row=2, max_row=3)
            cats = Reference(ws_rep, min_col=1, min_row=2, max_row=3)
            bar.add_data(data, titles_from_data=False)
            bar.set_categories(cats)
            bar.y_axis.title = "Amount"
            bar.y_axis.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            bar.width = 18
            bar.height = 10
            # Show values on bars
            bar.dataLabels = DataLabelList()
            bar.dataLabels.showVal = True
            bar.legend = None
            ws_rep.add_chart(bar, "F2")

            # Pie: Expense Categories based on breakdown table (exclude totals row)
            if 'start_row' in locals():
                pie = PieChart()
                pie.title = "Expense Categories"
                # end_row was defined before totals row append
                if 'end_row' in locals() and end_row >= start_row:
                    pdata = Reference(ws_rep, min_col=2, min_row=start_row, max_row=end_row)
                    pcats = Reference(ws_rep, min_col=1, min_row=start_row, max_row=end_row)
                    pie.add_data(pdata, titles_from_data=False)
                    pie.set_categories(pcats)
                    pie.width = 18
                    pie.height = 10
                    # Show percentages on slices
                    pie.dataLabels = DataLabelList()
                    pie.dataLabels.showPercent = True
                    pie.dataLabels.showLeaderLines = True
                    pie.legend.position = 'r'
                    # Apply a consistent color palette per slice
                    palette = [
                        "4CAF50", "2196F3", "FF9800", "9C27B0", "00BCD4",
                        "8BC34A", "FFC107", "E91E63", "795548", "607D8B",
                    ]
                    if pie.series:
                        # Build and color data points explicitly; 'Series.points' doesn't exist in openpyxl.
                        s = pie.series[0]
                        count = end_row - start_row + 1
                        s.data_points = []
                        for i in range(count):
                            dp = DataPoint(idx=i)
                            dp.graphicalProperties.solidFill = palette[i % len(palette)]
                            s.data_points.append(dp)
                    ws_rep.add_chart(pie, "F16")

            # Conditional formatting: highlight categories > 20% of total expenses
            try:
                if 'end_row' in locals() and end_row >= start_row:
                    high_fill = PatternFill("solid", fgColor="FFECEB")
                    rng = f"D{start_row}:D{end_row}"
                    rule = CellIsRule(operator='greaterThan', formula=['0.2'], stopIfTrue=False, fill=high_fill)
                    ws_rep.conditional_formatting.add(rng, rule)
            except Exception:
                pass

            # Auto-fit Report columns after charts
            self._excel_autofit_columns(ws_rep)

            wb.save(path)
        except Exception as exc:
            messagebox.showerror("Failed to export Excel", str(exc))
            return
        self.status.set(f"Excel saved: {os.path.basename(path)}")

    def _excel_autofit_columns(self, ws) -> None:
        try:
            max_widths: dict[int, int] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        length = 0
                    else:
                        text = str(cell.value)
                        length = len(text)
                    col_idx = cell.column if isinstance(cell.column, int) else cell.col_idx
                    max_widths[col_idx] = max(max_widths.get(col_idx, 0), min(60, length + 2))
            for col_idx, width in max_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = max(ws.column_dimensions[get_column_letter(col_idx)].width or 10, width)
        except Exception:
            # best effort only
            pass

    # Views
    def refresh_income_view(self) -> None:
        self.tv_income.delete(*self.tv_income.get_children())
        for inc in self.bm.incomes:
            self.tv_income.insert("", "end", values=(inc.name, f"{inc.amount:,.2f}", self._format_day_display(inc.date)))

    def refresh_expense_view(self) -> None:
        self.tv_expense.delete(*self.tv_expense.get_children())
        for exp in self.bm.expenses:
            self.tv_expense.insert("", "end", values=(exp.name, exp.category, f"{exp.amount:,.2f}", self._format_day_display(exp.date)))

    def update_report(self) -> None:
        self.bm.month = self.var_month.get().strip() or self.bm.month
        inc = self.bm.total_income()
        exp = self.bm.total_expenses()
        net = self.bm.net()
        pm = self.bm.profit_margin()
        self.lbl_income.config(text=f"Total Income: ${inc:,.2f}")
        self.lbl_expenses.config(text=f"Total Expenses: ${exp:,.2f}")
        self.lbl_net.config(text=f"Net (Profit): ${net:,.2f}")
        self.lbl_margin.config(text=f"Profit Margin: {pm:.2f}%")
        self.tv_breakdown.delete(*self.tv_breakdown.get_children())
        by_cat = self.bm.expenses_by_category()
        p_inc = self.bm.expense_percentages_by_category("income")
        p_exp = self.bm.expense_percentages_by_category("expenses")
        for cat in sorted(by_cat):
            amt = by_cat[cat]
            self.tv_breakdown.insert("", "end", values=(cat, f"{amt:,.2f}", f"{p_inc.get(cat, 0.0):.2f}%", f"{p_exp.get(cat, 0.0):.2f}%"))
        # Draw charts
        self._redraw_charts()

    def _redraw_charts(self) -> None:
        inc = self.bm.total_income()
        exp = self.bm.total_expenses()
        by_cat = self.bm.expenses_by_category()
        self._draw_income_expense_chart(inc, exp)
        self._draw_category_pie(by_cat)

    def _draw_income_expense_chart(self, income: float, expenses: float) -> None:
        c = self.canvas_income_expense
        c.delete("all")
        try:
            w = max(1, int(c.winfo_width()))
            h = max(1, int(c.winfo_height()))
        except Exception:
            return
        pad = 24
        usable_h = max(1, h - 2 * pad)
        usable_w = max(1, w - 3 * pad)
        bar_w = usable_w // 2
        max_val = max(income, expenses, 1.0)
        scale = usable_h / max_val
        # Bars X positions
        x1 = pad
        x2 = pad * 2 + bar_w
        # Heights
        h1 = int(income * scale)
        h2 = int(expenses * scale)
        y_base = h - pad
        # Draw axes baseline
        c.create_line(pad - 6, y_base, w - pad + 6, y_base, fill="#ccc")
        # Draw bars
        c.create_rectangle(x1, y_base - h1, x1 + bar_w, y_base, fill="#4caf50", outline="")
        c.create_rectangle(x2, y_base - h2, x2 + bar_w, y_base, fill="#f44336", outline="")
        # Labels
        c.create_text(x1 + bar_w / 2, y_base + 12, text="Income", fill="#333")
        c.create_text(x2 + bar_w / 2, y_base + 12, text="Expenses", fill="#333")
        c.create_text(x1 + bar_w / 2, y_base - h1 - 12, text=f"${income:,.0f}", fill="#222")
        c.create_text(x2 + bar_w / 2, y_base - h2 - 12, text=f"${expenses:,.0f}", fill="#222")

    def _draw_category_pie(self, by_cat: dict[str, float]) -> None:
        c = self.canvas_categories
        c.delete("all")
        try:
            w = max(1, int(c.winfo_width()))
            h = max(1, int(c.winfo_height()))
        except Exception:
            return
        pad = 16
        r = max(10, min(w, h) // 2 - pad)
        cx, cy = w // 2, h // 2
        total = sum(v for v in by_cat.values() if v > 0)
        if total <= 0:
            c.create_text(cx, cy, text="No expenses", fill="#666")
            return
        # Sort categories by amount desc
        items = sorted([(k, v) for k, v in by_cat.items() if v > 0], key=lambda kv: kv[1], reverse=True)
        bbox = (cx - r, cy - r, cx + r, cy + r)
        colors = [
            "#4caf50", "#2196f3", "#ff9800", "#9c27b0", "#00bcd4",
            "#8bc34a", "#ffc107", "#e91e63", "#795548", "#607d8b",
        ]
        angle = 0.0
        for i, (name, val) in enumerate(items):
            frac = val / total
            extent = frac * 360.0
            color = colors[i % len(colors)]
            c.create_arc(bbox, start=angle, extent=extent, fill=color, outline="white")
            # Label slices >= 5%
            if frac >= 0.05:
                mid = angle + extent / 2.0
                rad = _math.radians(mid)
                lx = cx + int(_math.cos(rad) * (r * 0.6))
                ly = cy - int(_math.sin(rad) * (r * 0.6))
                c.create_text(lx, ly, text=f"{name} {frac*100:.0f}%", fill="#222")
            angle += extent

    def _schedule_redraw(self) -> None:
        # Debounce rapid resizes
        try:
            if self._chart_redraw_after:
                self.after_cancel(self._chart_redraw_after)
        except Exception:
            pass
        self._chart_redraw_after = self.after(120, self._redraw_charts)

    # Preferences
    def _load_prefs(self) -> None:
        if not self._prefs_path:
            return
        try:
            if self._prefs_path.exists():
                data = json.loads(self._prefs_path.read_text(encoding="utf-8"))
                month = (data.get("month") or "").strip()
                if month:
                    self.var_month.set(month)
                    self.bm.month = month
        except Exception:
            # ignore prefs errors silently
            pass

    def _save_prefs(self) -> None:
        if not self._prefs_path:
            return
        try:
            data = {"month": (self.var_month.get() or "").strip()}
            self._prefs_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _confirm_discard_changes(self) -> bool:
        if self.bm.incomes or self.bm.expenses:
            return messagebox.askyesno("Discard changes", "Discard current entries?")
        return True

    def _format_day_display(self, ds: str | None) -> str:
        """Return display like '31 (Sun)' for YYYY-MM-DD; blank for YYYY-MM or None."""
        if not ds:
            return ""
        try:
            if len(ds) == 10:
                d = _dt.datetime.strptime(ds, "%Y-%m-%d").date()
                return f"{d.day} ({d.strftime('%a')})"
            return ""
        except Exception:
            return ds

    def _normalize_date_input(self, s: str) -> tuple[str | None, str | None]:
        """Normalize user-entered date:
        - Accept YYYY-MM, YYYY-MM-DD, or day-of-month (DD)
        - DD uses Month field or today's month
        Returns (normalized, error_message).
        """
        s = s.strip()
        if not s:
            m = (self.var_month.get() or "").strip()
            if _valid_ym(m):
                return m, None
            return None, None
        if _valid_ym(s) or _valid_ymd(s):
            return s, None
        if s.isdigit():
            try:
                day = int(s)
            except ValueError:
                return None, "Enter date as YYYY-MM, YYYY-MM-DD, or day-of-month (DD)."
            if not (1 <= day <= 31):
                return None, "Day must be between 1 and 31."
            base = (self.var_month.get() or "").strip()
            if not _valid_ym(base):
                base = _dt.date.today().strftime("%Y-%m")
            y = int(base[:4])
            m = int(base[5:7])
            try:
                d = _dt.date(y, m, day)
            except ValueError:
                return None, f"Day {day:02d} is not valid for {base}."
            return d.strftime("%Y-%m-%d"), None
        return None, "Enter date as YYYY-MM, YYYY-MM-DD, or day-of-month (DD)."

    def _infer_month_from_entries(self) -> str | None:
        counts: dict[str, int] = {}
        def add_date(ds: str | None) -> None:
            if not ds:
                return
            ym: str | None = None
            if _valid_ymd(ds):
                ym = ds[:7]
            elif _valid_ym(ds):
                ym = ds
            if ym:
                counts[ym] = counts.get(ym, 0) + 1
        for inc in self.bm.incomes:
            add_date(getattr(inc, 'date', None))
        for exp in self.bm.expenses:
            add_date(getattr(exp, 'date', None))
        if not counts:
            return None
        # Pick the month with max occurrences; tie-breaker uses lexicographic order (earlier month first)
        return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]


def main() -> int:
    app = BudgetApp()
    try:
        app.var_month.set(_dt.date.today().strftime("%Y-%m"))
    except Exception:
        pass
    app.mainloop()
    return 0


class _CalendarPopup(tk.Toplevel):
    def __init__(self, master: tk.Misc, year: int, month: int) -> None:
        super().__init__(master)
        self.title("Pick a date")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._selected: str | None = None
        self._year = year
        self._month = month
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build(self) -> None:
        hdr = ttk.Frame(self, padding=8)
        ttk.Button(hdr, text="◀", width=3, command=self._prev_month).grid(row=0, column=0)
        ttk.Label(hdr, text=f"{self._year}-{self._month:02d}", width=10, anchor="center").grid(row=0, column=1, padx=8)
        ttk.Button(hdr, text="▶", width=3, command=self._next_month).grid(row=0, column=2)
        hdr.grid(row=0, column=0, sticky="ew")

        self.grid_days = ttk.Frame(self, padding=(8, 0))
        self.grid_days.grid(row=1, column=0)
        self._render_days()

        btns = ttk.Frame(self, padding=8)
        ttk.Button(btns, text="Cancel", command=self._on_cancel).grid(row=0, column=0, padx=4)
        btns.grid(row=2, column=0)

    def _render_days(self) -> None:
        for w in list(self.grid_days.winfo_children()):
            w.destroy()
        for i, wd in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            ttk.Label(self.grid_days, text=wd, width=4, anchor="center").grid(row=0, column=i, padx=1, pady=2)
        cal = _cal.Calendar(firstweekday=0)
        row = 1
        for week in cal.monthdayscalendar(self._year, self._month):
            for col, day in enumerate(week):
                if day == 0:
                    ttk.Label(self.grid_days, text="", width=4).grid(row=row, column=col, padx=1, pady=1)
                else:
                    def make_cmd(d=day):
                        return lambda: self._on_pick(d)
                    ttk.Button(self.grid_days, text=f"{day:02d}", width=4, command=make_cmd()).grid(row=row, column=col, padx=1, pady=1)
            row += 1

    def _on_pick(self, day: int) -> None:
        d = _dt.date(self._year, self._month, day)
        self._selected = d.strftime("%Y-%m-%d")
        self.destroy()

    def _on_cancel(self) -> None:
        self._selected = None
        self.destroy()

    def _prev_month(self) -> None:
        if self._month == 1:
            self._month = 12
            self._year -= 1
        else:
            self._month -= 1
        self._render_days()

    def _next_month(self) -> None:
        if self._month == 12:
            self._month = 1
            self._year += 1
        else:
            self._month += 1
        self._render_days()

    def show(self) -> str | None:
        self.wait_window()
        return self._selected


class _CategoryPicker(tk.Toplevel):
    def __init__(self, master: tk.Misc, categories: list[str]) -> None:
        super().__init__(master)
        self.title("Pick a category")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._selected: str | None = None
        self._cats = categories
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)

    def _build(self) -> None:
        frm = ttk.Frame(self, padding=8)
        frm.grid(row=0, column=0)
        lb = tk.Listbox(frm, height=min(12, max(6, len(self._cats))), exportselection=False)
        for c in self._cats:
            lb.insert(tk.END, c)
        lb.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(frm, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        sb.grid(row=0, column=1, sticky="ns")

        btns = ttk.Frame(self, padding=8)
        ttk.Button(btns, text="OK", command=lambda: self._on_ok(lb)).grid(row=0, column=0, padx=4)
        ttk.Button(btns, text="Cancel", command=self._on_cancel).grid(row=0, column=1, padx=4)
        btns.grid(row=1, column=0)

    def _on_ok(self, lb: tk.Listbox) -> None:
        sel = lb.curselection()
        if sel:
            self._selected = lb.get(sel[0])
        self.destroy()

    def _on_cancel(self) -> None:
        self._selected = None
        self.destroy()

    def show(self) -> str | None:
        self.wait_window()
        return self._selected


if __name__ == "__main__":
    raise SystemExit(main())
