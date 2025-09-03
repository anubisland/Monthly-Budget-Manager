import json
import io
import sys
from datetime import datetime
from pathlib import Path

import toga
from .model import totals, expenses_by_category, parse_amount


CATEGORIES = [
	"Food", "Rent", "Fuel", "Electricity", "Internet", "Water", "Transport",
	"Healthcare", "Entertainment", "Education", "Clothing", "Savings",
	"Debt", "Subscriptions", "Gifts", "Misc", "Uncategorized"
]


class BudgetMobile(toga.App):
	def startup(self):
		# Data stores
		self.incomes = []
		self.expenses = []
		# File meta
		today = datetime.now()
		self.meta_year = today.year
		self.meta_month = today.month

		# Small layout helper for labeled rows
		def form_row(label: str, widget: toga.Widget) -> toga.Box:
			lbl = toga.Label(label, style=toga.style.Pack(width=110, padding_bottom=6))
			box = toga.Box(style=toga.style.Pack(direction=toga.style.pack.ROW, padding_bottom=6))
			# Toga 0.5: update style using keyword args rather than passing a Pack instance
			widget.style.update(flex=1)
			box.add(lbl)
			box.add(widget)
			return box

		# Income tab
		self.i_name = toga.TextInput(placeholder="e.g. Salary", style=toga.style.Pack())
		self.i_amount = toga.TextInput(placeholder="e.g. 1200.00", style=toga.style.Pack())
		self.i_day = toga.Selection(items=[str(i).zfill(2) for i in range(1, 32)], style=toga.style.Pack(width=120))
		self.i_today = toga.Button("Today", on_press=self.use_today_income_day, style=toga.style.Pack(width=100))
		row_date_income = toga.Box(style=toga.style.Pack(direction=toga.style.pack.ROW, padding_bottom=6))
		row_date_income.add(toga.Label("Day", style=toga.style.Pack(width=110, padding_right=6)))
		row_date_income.add(self.i_day)
		row_date_income.add(self.i_today)
		self.i_add = toga.Button("Add income", on_press=self.add_income, style=toga.style.Pack(padding_top=4, padding_bottom=8))
		self.i_table = toga.Table(headings=["Name", "Amount", "Date"], data=[], style=toga.style.Pack(flex=1, height=300))
		income_box = toga.Box(style=toga.style.Pack(direction=toga.style.pack.COLUMN, padding=8))
		income_box.add(toga.Label("Income", style=toga.style.Pack(padding_bottom=8)))
		income_box.add(form_row("Name", self.i_name))
		income_box.add(form_row("Amount", self.i_amount))
		income_box.add(row_date_income)
		income_box.add(self.i_add)
		income_box.add(self.i_table)

		# Expenses tab
		self.e_name = toga.TextInput(placeholder="e.g. Groceries", style=toga.style.Pack())
		self.e_category_select = toga.Selection(items=CATEGORIES, style=toga.style.Pack())
		self.e_category = toga.TextInput(placeholder="Custom category (optional)", style=toga.style.Pack())
		self.e_amount = toga.TextInput(placeholder="e.g. 50.00", style=toga.style.Pack())
		self.e_day = toga.Selection(items=[str(i).zfill(2) for i in range(1, 32)], style=toga.style.Pack(width=120))
		self.e_today = toga.Button("Today", on_press=self.use_today_expense_day, style=toga.style.Pack(width=100))
		row_date_expense = toga.Box(style=toga.style.Pack(direction=toga.style.pack.ROW, padding_bottom=6))
		row_date_expense.add(toga.Label("Day", style=toga.style.Pack(width=110, padding_right=6)))
		row_date_expense.add(self.e_day)
		row_date_expense.add(self.e_today)
		self.e_add = toga.Button("Add expense", on_press=self.add_expense, style=toga.style.Pack(padding_top=4, padding_bottom=8))
		self.e_table = toga.Table(headings=["Name", "Category", "Amount", "Date"], data=[], style=toga.style.Pack(flex=1, height=300))
		expense_box = toga.Box(style=toga.style.Pack(direction=toga.style.pack.COLUMN, padding=8))
		expense_box.add(toga.Label("Expenses", style=toga.style.Pack(padding_bottom=8)))
		expense_box.add(form_row("Name", self.e_name))
		expense_box.add(form_row("Category", self.e_category_select))
		expense_box.add(form_row("Custom", self.e_category))
		expense_box.add(form_row("Amount", self.e_amount))
		expense_box.add(row_date_expense)
		expense_box.add(self.e_add)
		expense_box.add(self.e_table)

		# Report tab
		# Date controls
		self.year_input = toga.TextInput(placeholder="YYYY", style=toga.style.Pack())
		self.month_select = toga.Selection(items=[str(i).zfill(2) for i in range(1, 13)], style=toga.style.Pack(width=120))
		self.today_btn = toga.Button("Today", on_press=self.use_today, style=toga.style.Pack(width=100))

		self.r_income_total = toga.Label("")
		self.r_expense_total = toga.Label("")
		self.r_profit = toga.Label("")
		self.r_margin = toga.Label("")
		self.r_categories = toga.Table(headings=["Category", "Amount", "Percent"], data=[], style=toga.style.Pack(flex=1, height=300))
		self.status_label = toga.Label("")

		report_box = toga.Box(style=toga.style.Pack(direction=toga.style.pack.COLUMN, padding=8))
		for w in (
			toga.Label("Report", style=toga.style.Pack(padding_bottom=8)),
			toga.Label("Period (YYYY-MM)"),
			# period row
			toga.Box(children=[
				self.year_input,
				self.month_select,
				self.today_btn,
			], style=toga.style.Pack(direction=toga.style.pack.ROW, padding_bottom=8, alignment=toga.style.pack.LEFT)),
			toga.Label("Totals", style=toga.style.Pack(padding_top=4)),
			self.r_income_total,
			self.r_expense_total,
			self.r_profit,
			self.r_margin,
			toga.Label("Expense Categories", style=toga.style.Pack(padding_top=6)),
			self.r_categories,
			self.status_label,
		):
			report_box.add(w)

		# Content areas
		self.income_box = income_box
		self.expense_box = expense_box
		self.report_box = report_box
		self.content = toga.Box(style=toga.style.Pack(direction=toga.style.pack.COLUMN, flex=1))
		self.content.add(self.income_box)

		# Bottom navigation with icons
		self.nav_income = toga.Button("💰 Income", on_press=lambda b: self.switch_section("Income"), style=toga.style.Pack(flex=1, padding=6))
		self.nav_expenses = toga.Button("🧾 Expenses", on_press=lambda b: self.switch_section("Expenses"), style=toga.style.Pack(flex=1, padding=6))
		self.nav_report = toga.Button("📊 Report", on_press=lambda b: self.switch_section("Report"), style=toga.style.Pack(flex=1, padding=6))
		self.navbar = toga.Box(children=[self.nav_income, self.nav_expenses, self.nav_report], style=toga.style.Pack(direction=toga.style.pack.ROW, padding=(4, 8)))

		root = toga.Box(style=toga.style.Pack(direction=toga.style.pack.COLUMN))
		root.add(self.content)
		root.add(self.navbar)

		self.main_window = toga.MainWindow(title=self.formal_name)
		self.main_window.content = root
		# Overflow menu commands (3-dots)
		self.cmd_new = toga.Command(self.on_new, text="New", tooltip="Start a new monthly report")
		self.cmd_save = toga.Command(self.on_save, text="Save", tooltip="Save to JSON")
		self.cmd_open = toga.Command(self.on_open, text="Open", tooltip="Open from JSON")
		self.cmd_export = toga.Command(self.on_export, text="Export", tooltip="Export to Excel")
		# Add to application commands so they appear in the 3-dots overflow menu (alongside About)
		self.commands.add(self.cmd_new, self.cmd_save, self.cmd_open, self.cmd_export)
		self.main_window.show()
		# Initialize date controls
		self._sync_date_controls()
		self.update_report()
		# Default day selections to today
		today_day = datetime.now().day
		self.i_day.value = str(today_day).zfill(2)
		self.e_day.value = str(today_day).zfill(2)
		# Highlight Income tab by default
		self._highlight_nav("Income")

	# Actions
	def _compose_date(self, day: int | None) -> str | None:
		if not day:
			return None
		return f"{self.meta_year}-{str(self.meta_month).zfill(2)}-{str(day).zfill(2)}"

	def add_income(self, button):
		name = (self.i_name.value or "").strip()
		amt = parse_amount(self.i_amount.value)
		day = None
		try:
			day = int((self.i_day.value or "").strip())
		except Exception:
			day = None
		date_str = self._compose_date(day)
		if name or amt:
			rec = {"name": name, "amount": amt}
			if date_str:
				rec["date"] = date_str
			self.incomes.append(rec)
			self.i_table.data.append([name, f"{amt:.2f}", date_str or ""])
			self.i_name.value = ""
			self.i_amount.value = ""
			self.update_report()

	def add_expense(self, button):
		name = (self.e_name.value or "").strip()
		custom_cat = (self.e_category.value or "").strip()
		cat = custom_cat or (self.e_category_select.value or "Uncategorized").strip()
		amt = parse_amount(self.e_amount.value)
		day = None
		try:
			day = int((self.e_day.value or "").strip())
		except Exception:
			day = None
		date_str = self._compose_date(day)
		if name or cat or amt:
			rec = {"name": name, "category": cat, "amount": amt}
			if date_str:
				rec["date"] = date_str
			self.expenses.append(rec)
			self.e_table.data.append([name, cat, f"{amt:.2f}", date_str or ""])
			self.e_name.value = ""
			self.e_category.value = ""
			self.e_amount.value = ""
			self.update_report()

	def update_report(self):
		t = totals(self.incomes, self.expenses)
		self.r_income_total.text = f"Income Total: {t['income_total']:.2f}"
		self.r_expense_total.text = f"Expense Total: {t['expense_total']:.2f}"
		self.r_profit.text = f"Profit: {t['profit']:.2f}"
		self.r_margin.text = f"Profit Margin: {t['profit_margin']:.2f}%"
		# Categories table
		cats = expenses_by_category(self.expenses)
		self.r_categories.data.clear()
		for row in cats:
			self.r_categories.data.append([
				row["category"], f"{row['amount']:.2f}", f"{row['percent']:.1f}%",
			])

	def _highlight_nav(self, section: str):
		# Simple highlight: disable active button
		self.nav_income.enabled = section != "Income"
		self.nav_expenses.enabled = section != "Expenses"
		self.nav_report.enabled = section != "Report"

	def switch_section(self, section: str):
		# Swap displayed content based on requested section
		for child in list(self.content.children):
			self.content.remove(child)
		if section == "Income":
			self.content.add(self.income_box)
		elif section == "Expenses":
			self.content.add(self.expense_box)
		else:
			self.content.add(self.report_box)
		self._highlight_nav(section)

	# Date helpers
	def _is_mobile(self) -> bool:
		# Android reports 'android'; iOS may report 'ios'
		plat = sys.platform.lower()
		return plat.startswith("android") or plat == "ios"

	def _sync_date_controls(self):
		self.year_input.value = str(self.meta_year)
		self.month_select.value = str(self.meta_month).zfill(2)

	def _read_date_controls(self):
		try:
			y = int((self.year_input.value or '').strip())
		except Exception:
			y = self.meta_year
		try:
			m = int((self.month_select.value or '').strip())
		except Exception:
			m = self.meta_month
		m = max(1, min(12, m))
		if y < 1900 or y > 3000:
			y = self.meta_year
		self.meta_year, self.meta_month = y, m
		self._sync_date_controls()

	def use_today(self, button):
		now = datetime.now()
		self.meta_year, self.meta_month = now.year, now.month
		self._sync_date_controls()

	# Persistence
	def _serialize(self):
		self._read_date_controls()
		return {
			"meta": {
				"year": self.meta_year,
				"month": self.meta_month,
				"saved_at": datetime.now().isoformat(timespec="seconds"),
			},
			"incomes": self.incomes,
			"expenses": self.expenses,
		}

	def _deserialize(self, data: dict):
		meta = data.get("meta", {})
		self.meta_year = int(meta.get("year", self.meta_year))
		self.meta_month = int(meta.get("month", self.meta_month))
		self.incomes = list(data.get("incomes", []))
		self.expenses = list(data.get("expenses", []))
		# Rebuild tables
		self.i_table.data.clear()
		for r in self.incomes:
			self.i_table.data.append([
				r.get("name", ""),
				f"{parse_amount(r.get('amount', 0)):.2f}",
				r.get("date", ""),
			])
		self.e_table.data.clear()
		for r in self.expenses:
			self.e_table.data.append([
				r.get("name", ""),
				r.get("category", ""),
				f"{parse_amount(r.get('amount', 0)):.2f}",
				r.get("date", ""),
			])
		self._sync_date_controls()
		self.update_report()

	def use_today_income_day(self, button):
		self.i_day.value = str(datetime.now().day).zfill(2)

	def use_today_expense_day(self, button):
		self.e_day.value = str(datetime.now().day).zfill(2)

	def _default_filename(self, ext: str) -> str:
		return f"budget-{self.meta_year}-{str(self.meta_month).zfill(2)}.{ext}"

	def _safe_app_path(self, filename: str) -> Path:
		try:
			base = Path(self.paths.data)
		except Exception:
			base = Path.home()
		base.mkdir(parents=True, exist_ok=True)
		return base / filename

	def _first_selection(self, sel):
		"""Normalize dialog return to a single selection (str or Document-like).
		Accepts str, list/tuple of items, or an object with open()/path attributes.
		"""
		if sel is None:
			return None
		if isinstance(sel, (list, tuple)):
			return sel[0] if sel else None
		return sel

	async def _confirm(self, title: str, message: str) -> bool:
		try:
			# Toga confirm dialog returns True/False
			return await self.main_window.confirm_dialog(title, message)
		except Exception:
			return False

	async def on_save(self, button):
		try:
			# Use a native save dialog; require explicit selection
			sel = await self.main_window.save_file_dialog(
				"Save Budget JSON",
				suggested_filename=self._default_filename("json"),
				file_types=None,
			)
			sel = self._first_selection(sel)
			if not sel:
				# Auto fallback to app storage default path
				sel = str(self._safe_app_path(self._default_filename("json")))
			data = self._serialize()
			# If the selection is a Document-like object, use its open() method
			if hasattr(sel, "open"):
				with sel.open("w", encoding="utf-8") as f:
					json.dump(data, f, ensure_ascii=False, indent=2)
				self.status_label.text = "Saved"
			else:
				path = str(sel)
				with open(path, "w", encoding="utf-8") as f:
					json.dump(data, f, ensure_ascii=False, indent=2)
				self.status_label.text = f"Saved: {path}"
		except Exception as e:
			self.status_label.text = f"Save failed: {e}"

	async def on_open(self, button):
		try:
			# Use a native open dialog; require explicit selection
			sel = await self.main_window.open_file_dialog(
				"Open Budget JSON",
				multiselect=False,
				file_types=None,
			)
			sel = self._first_selection(sel)
			if not sel:
				# Auto fallback to default file in app storage
				sel = str(self._safe_app_path(self._default_filename("json")))
			if hasattr(sel, "open"):
				with sel.open("r", encoding="utf-8") as f:
					data = json.load(f)
			else:
				path = str(sel)
				with open(path, "r", encoding="utf-8") as f:
					data = json.load(f)
			self._deserialize(data)
			self.status_label.text = "Opened"
		except FileNotFoundError:
			self.status_label.text = "Open failed: file not found"
		except Exception as e:
			self.status_label.text = f"Open failed: {e}"

	async def on_export(self, button):
		try:
			# Try XlsxWriter first (pure-Python), then openpyxl
			engine = None
			try:
				import xlsxwriter  # type: ignore
				engine = "xlsxwriter"
			except Exception:
				pass
			if engine is None:
				try:
					from openpyxl import Workbook  # type: ignore
					from openpyxl.utils import get_column_letter  # type: ignore
					from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
					engine = "openpyxl"
				except Exception:
					engine = None
			if engine is None:
				self.status_label.text = "Export failed: No Excel engine (install XlsxWriter or openpyxl)"
				return

			# Ask user where to save; require explicit selection
			sel = await self.main_window.save_file_dialog(
				"Export to Excel",
				suggested_filename=self._default_filename("xlsx"),
				file_types=None,
			)
			sel = self._first_selection(sel)
			if not sel:
				# Auto fallback to app storage
				sel = str(self._safe_app_path(self._default_filename("xlsx")))

			if engine == "xlsxwriter":
				# Build workbook with XlsxWriter
				import xlsxwriter  # type: ignore
				buffer = io.BytesIO()
				wb = xlsxwriter.Workbook(buffer, {"in_memory": True})
				fmt_bold = wb.add_format({"bold": True})
				fmt_hdr = wb.add_format({"bold": True, "bg_color": "#DDDDDD", "align": "center"})
				# Summary
				sum_ws = wb.add_worksheet("Summary")
				sum_ws.write_row(0, 0, ["Year", self.meta_year], fmt_bold)
				sum_ws.write_row(1, 0, ["Month", self.meta_month])
				sum_ws.write(2, 0, "")
				t = totals(self.incomes, self.expenses)
				sum_ws.write_row(3, 0, ["Income Total", t["income_total"]])
				sum_ws.write_row(4, 0, ["Expense Total", t["expense_total"]])
				sum_ws.write_row(5, 0, ["Profit", t["profit"]])
				sum_ws.write_row(6, 0, ["Profit Margin %", t["profit_margin"]])
				sum_ws.set_column(0, 1, 20)
				# Incomes
				inc = wb.add_worksheet("Incomes")
				inc.write_row(0, 0, ["Name", "Amount", "Date"], fmt_hdr)
				rowi = 1
				for r in self.incomes:
					inc.write_row(rowi, 0, [r.get("name", ""), parse_amount(r.get("amount", 0)), r.get("date", "")])
					rowi += 1
				inc.set_column(0, 0, 30)
				inc.set_column(1, 1, 15)
				inc.set_column(2, 2, 15)
				# Expenses
				exp = wb.add_worksheet("Expenses")
				exp.write_row(0, 0, ["Name", "Category", "Amount", "Date"], fmt_hdr)
				rowe = 1
				for r in self.expenses:
					exp.write_row(rowe, 0, [r.get("name", ""), r.get("category", ""), parse_amount(r.get("amount", 0)), r.get("date", "")])
					rowe += 1
				exp.set_column(0, 0, 30)
				exp.set_column(1, 1, 20)
				exp.set_column(2, 2, 15)
				exp.set_column(3, 3, 15)
				# Categories
				cat = wb.add_worksheet("Categories")
				cat.write_row(0, 0, ["Category", "Amount", "Percent"], fmt_hdr)
				rowc = 1
				for row in expenses_by_category(self.expenses):
					cat.write_row(rowc, 0, [row["category"], row["amount"], row["percent"]])
					rowc += 1
				cat.set_column(0, 2, 20)
				wb.close()
				data_bytes = buffer.getvalue()
			else:
				# Build workbook with openpyxl
				from openpyxl import Workbook
				from openpyxl.utils import get_column_letter
				from openpyxl.styles import Font, Alignment, PatternFill
				wb = Workbook()
				ws = wb.active
				ws.title = "Summary"
				ws.append(["Year", self.meta_year])
				ws.append(["Month", self.meta_month])
				ws.append([])
				t = totals(self.incomes, self.expenses)
				ws.append(["Income Total", t["income_total"]])
				ws.append(["Expense Total", t["expense_total"]])
				ws.append(["Profit", t["profit"]])
				ws.append(["Profit Margin %", t["profit_margin"]])
				for col in range(1, 3):
					ws.column_dimensions[get_column_letter(col)].width = 20
				for cell in ws[1]:
					cell.font = Font(bold=True)
				inc = wb.create_sheet("Incomes")
				inc.append(["Name", "Amount", "Date"])
				for c in inc[1]:
					c.font = Font(bold=True)
					c.fill = PatternFill("solid", fgColor="DDDDDD")
					c.alignment = Alignment(horizontal="center")
				for r in self.incomes:
					inc.append([r.get("name", ""), parse_amount(r.get("amount", 0)), r.get("date", "")])
				inc.column_dimensions['A'].width = 30
				inc.column_dimensions['B'].width = 15
				inc.column_dimensions['C'].width = 15
				exp = wb.create_sheet("Expenses")
				exp.append(["Name", "Category", "Amount", "Date"])
				for c in exp[1]:
					c.font = Font(bold=True)
					c.fill = PatternFill("solid", fgColor="DDDDDD")
					c.alignment = Alignment(horizontal="center")
				for r in self.expenses:
					exp.append([r.get("name", ""), r.get("category", ""), parse_amount(r.get("amount", 0)), r.get("date", "")])
				exp.column_dimensions['A'].width = 30
				exp.column_dimensions['B'].width = 20
				exp.column_dimensions['C'].width = 15
				exp.column_dimensions['D'].width = 15
				cat = wb.create_sheet("Categories")
				cat.append(["Category", "Amount", "Percent"])
				for c in cat[1]:
					c.font = Font(bold=True)
					c.fill = PatternFill("solid", fgColor="DDDDDD")
					c.alignment = Alignment(horizontal="center")
				for row in expenses_by_category(self.expenses):
					cat.append([row["category"], row["amount"], row["percent"]])
				for col in ('A','B','C'):
					cat.column_dimensions[col].width = 20
				buffer = io.BytesIO()
				wb.save(buffer)
				data_bytes = buffer.getvalue()

			# Write bytes to target selection or fallback path
			if hasattr(sel, "open"):
				with sel.open("wb") as f:
					f.write(data_bytes)
				self.status_label.text = "Exported"
			else:
				path = str(sel) if sel else str(self._safe_app_path(self._default_filename("xlsx")))
				with open(path, "wb") as f:
					f.write(data_bytes)
				self.status_label.text = f"Exported: {path}"
		except Exception as e:
			self.status_label.text = f"Export failed: {e}"

	def on_new(self, button):
		# Reset to a new monthly report
		now = datetime.now()
		self.meta_year, self.meta_month = now.year, now.month
		self.incomes = []
		self.expenses = []
		self.i_table.data.clear()
		self.e_table.data.clear()
		self.i_name.value = ""
		self.i_amount.value = ""
		self.e_name.value = ""
		self.e_amount.value = ""
		self.e_category.value = ""
		self._sync_date_controls()
		self.i_day.value = str(now.day).zfill(2)
		self.e_day.value = str(now.day).zfill(2)
		self.update_report()
		self.status_label.text = "Started new monthly report"


def main():
	return BudgetMobile("Budget Manager", "com.example.budget_manager_mobile")
